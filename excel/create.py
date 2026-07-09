from openpyxl import Workbook,load_workbook

def create_excel(filename: str, data: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MySheet"
    for row in data:
        ws.append(row)
    wb.save(filename)
    wb.close()

def read_excel(filename: str) -> list[list[str]]:
    wb = load_workbook(filename)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    wb.close()
    return data
ss = read_excel("example.xlsx")
print(ss)